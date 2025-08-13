import pandas as pd
import re
from openpyxl import load_workbook

file = "Time-Table, FSC, Fall-2025.xlsx"

def normalize_color(fgColor):
    """Convert openpyxl fgColor to hex (#RRGGBB)."""
    if fgColor and fgColor.type == "rgb" and fgColor.rgb:
        return f"#{fgColor.rgb[-6:].upper()}"
    return None

def extract_color_batch_map(file_path, sheet_name):
    """Read first 4 rows but only valid legend cells for colors → batches."""
    wb = load_workbook(file_path, data_only=False)
    ws = wb[sheet_name]
    mapping = {}
    ignore_words = ["monday", "tuesday", "wednesday", "thursday", "friday", 
                    "room", "timetable", "time", "slot"]

    for row in ws.iter_rows(min_row=1, max_row=4):
        for cell in row:
            color = normalize_color(cell.fill.fgColor)
            if not color or color == "#FFFFFF":  # skip white/empty fills
                continue
            if cell.value and isinstance(cell.value, str):
                text = cell.value.strip()
                if any(word in text.lower() for word in ignore_words):
                    continue
                mapping[color] = text
    return mapping

def get_excel_column_mapping(ws, header_row_excel):
    """Create mapping from time slot strings to Excel column numbers."""
    time_pattern = re.compile(r"\d{2}:\d{2}-\d{2}:\d{2}")
    mapping = {}
    
    # Check all columns in the header row
    for col_num in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=header_row_excel, column=col_num).value
        if cell_value and isinstance(cell_value, str):
            cell_value_clean = str(cell_value).strip()
            if time_pattern.search(cell_value_clean):
                # Map both cleaned and original versions to handle whitespace differences
                mapping[cell_value_clean] = col_num
                mapping[str(cell_value)] = col_num  # Also map the original with any whitespace
    
    return mapping

def extract_section_from_course(course_name):
    """Extract section/class info from course name like 'PF (CS-E)' -> 'CS-E'"""
    if not course_name or course_name == "Free Slot":
        return None, course_name
    
    # Pattern to match section in parentheses like (CS-E), (CS-A), (SE), (CY), etc.
    section_pattern = r'\(([A-Z]{2,3}(?:-[A-Z])?)\)$'
    match = re.search(section_pattern, course_name)
    
    if match:
        section = match.group(1)
        # Remove the section part from course name
        clean_course_name = re.sub(section_pattern, '', course_name).strip()
        return section, clean_course_name
    
    return None, course_name

def reshape_timetable(df, day_name):
    # Store original for reference
    original_df = df.copy()
    
    # Drop empty rows & columns
    df = df.dropna(how="all").dropna(axis=1, how="all").reset_index(drop=True)

    # Find "Room" header row in cleaned DataFrame
    header_row_index = df.index[df.iloc[:, 0].astype(str).str.contains("Room", case=False, na=False)]
    if len(header_row_index) == 0:
        return pd.DataFrame(columns=["Day", "Course Name", "Class Time", "Room No", "Section", "Batch"])
    
    header_row_pandas = header_row_index[0]

    # Load the workbook for color information
    wb = load_workbook(file, data_only=False)
    ws = wb[day_name]
    
    # Find header row in Excel worksheet
    header_row_excel = None
    for row_num in range(1, 20):  # Check first 20 rows
        cell_value = ws.cell(row=row_num, column=1).value
        if cell_value and "room" in str(cell_value).lower():
            header_row_excel = row_num
            break
    
    if header_row_excel is None:
        return pd.DataFrame(columns=["Day", "Course Name", "Class Time", "Room No", "Section", "Batch"])
    
    # Get color to batch mapping
    color_batch_map = extract_color_batch_map(file, day_name)
    
    # Get mapping from time slots to Excel columns
    excel_col_mapping = get_excel_column_mapping(ws, header_row_excel)

    df.columns = df.iloc[header_row_pandas]
    df = df.iloc[header_row_pandas + 1:].reset_index(drop=True)

    # Identify time columns
    time_pattern = re.compile(r"\d{2}:\d{2}-\d{2}:\d{2}")
    time_columns = [col for col in df.columns if isinstance(col, str) and time_pattern.search(col)]

    results = []
    for r_idx, row in df.iterrows():
        room = row.get("Room", "Unknown")
        for col in time_columns:
            raw_course = str(row[col]) if pd.notna(row[col]) else "Free Slot"

            # Extract actual time if embedded in course name
            found_time = time_pattern.search(raw_course)
            if found_time:
                actual_time = found_time.group()
                course_name = raw_course.replace(actual_time, "").strip()
            else:
                actual_time = col
                course_name = raw_course

            # Extract section/class from course name
            section, clean_course_name = extract_section_from_course(course_name)

            # Calculate correct Excel position
            excel_row = header_row_excel + 1 + r_idx  # +1 to skip header, r_idx for data row
            excel_col = excel_col_mapping.get(col, None)
            if excel_col is None:
                # Try with stripped version
                excel_col = excel_col_mapping.get(col.strip(), None)
            if excel_col is None:
                # Try to find by pattern matching (fallback)
                col_clean = col.strip()
                for mapped_col, mapped_num in excel_col_mapping.items():
                    if mapped_col.strip() == col_clean:
                        excel_col = mapped_num
                        break
            
            # Get cell color from Excel
            if excel_col is not None:
                cell_color = normalize_color(ws.cell(row=excel_row, column=excel_col).fill.fgColor)
            else:
                cell_color = None
                print(f"Warning: Could not find Excel column for time slot '{col}'")
            
            # Only assign batch if we have an exact color match
            batch = color_batch_map.get(cell_color, None)

            results.append({
                "Day": day_name,
                "Course Name": clean_course_name,
                "Class Time": actual_time,
                "Room No": room,
                "Section": section,
                "Batch": batch,
            })
    
    # Add this at the end of reshape_timetable function, before return
    exclude_times = ["05:20-06:40", "06:45-08:05"]
    results_df = pd.DataFrame(results)
    results_df = results_df[~results_df["Class Time"].isin(exclude_times)]
    return results_df

# Load all sheet names
sheet_names = pd.ExcelFile(file).sheet_names
print(sheet_names)

time_table_data = {}

for names in sheet_names:
    if names != "Welcome":
        data = pd.read_excel(file, sheet_name=names)
        time_table_data[names] = data

# Process all days
event_tables = {day: reshape_timetable(time_table_data[day], day) for day in time_table_data}

# Example: Monday
print(event_tables["Monday"].head(30))

# Optionally save to Excel
#all_days_df = pd.concat(event_tables.values())
#all_days_df.to_excel("complete_timetable_with_batches.xlsx", index=False)