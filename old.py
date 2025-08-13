import pandas as pd
import re
from openpyxl import load_workbook

file = "Time-Table, FSC, Fall-2025.xlsx"

def normalize_color(fgColor):
    if fgColor and fgColor.type == "rgb" and fgColor.rgb:
        return f"#{fgColor.rgb[-6:].upper()}"
    return None

def normalize_time_str(time_str):
    """Standardize time format to HH:MM-HH:MM with no spaces"""
    if not time_str or not isinstance(time_str, str):
        return None
    time_str = re.sub(r'\s*[-–]\s*', '-', time_str.strip())  # normalize dash and remove spaces
    def pad_hour(match):
        hour, minute = match.groups()
        return f"{int(hour):02}:{minute}"
    time_str = re.sub(r'(\d{1,2}):(\d{2})', pad_hour, time_str)
    return time_str

def extract_color_batch_map(file_path, sheet_name):
    wb = load_workbook(file_path, data_only=False)
    ws = wb[sheet_name]
    mapping = {}
    ignore_words = ["monday", "tuesday", "wednesday", "thursday", "friday", 
                    "room", "timetable", "time", "slot"]
    for row in ws.iter_rows(min_row=1, max_row=4):
        for cell in row:
            color = normalize_color(cell.fill.fgColor)
            if not color or color == "#FFFFFF":
                continue
            if cell.value and isinstance(cell.value, str):
                text = cell.value.strip()
                if any(word in text.lower() for word in ignore_words):
                    continue
                mapping[color] = text
    return mapping

def parse_time_to_minutes(time_str):
    try:
        parts = time_str.split(':')
        hours = int(parts[0])
        minutes = int(parts[1])
        return hours * 60 + minutes
    except:
        return None

def time_ranges_overlap(time1, time2):
    if not time1 or not time2:
        return False
    try:
        start1, end1 = time1.split('-')
        start2, end2 = time2.split('-')
        start1_min = parse_time_to_minutes(start1)
        end1_min = parse_time_to_minutes(end1)
        start2_min = parse_time_to_minutes(start2)
        end2_min = parse_time_to_minutes(end2)
        if None in [start1_min, end1_min, start2_min, end2_min]:
            return False
        return not (end1_min <= start2_min or end2_min <= start1_min)
    except:
        return False

def create_excel_to_dataframe_mapping(ws, df_columns, header_row_excel):
    mapping = {}
    excel_time_cols = {}

    # Scan header row + a few rows below it (to catch merged cells)
    for col_num in range(1, ws.max_column + 1):
        for row_num in range(header_row_excel, header_row_excel + 3):
            cell_value = ws.cell(row=row_num, column=col_num).value
            if cell_value:
                cell_str = normalize_time_str(str(cell_value))
                if re.search(r'\d{2}:\d{2}-\d{2}:\d{2}', str(cell_str)):
                    excel_time_cols[col_num] = cell_str
                    break  # Stop after finding first valid time in this column

    # Map DataFrame columns to Excel columns
    for df_col in df_columns:
        if pd.isna(df_col) or str(df_col).strip() == 'nan':
            continue
        df_col_str = normalize_time_str(str(df_col))
        for excel_col_num, excel_col_str in excel_time_cols.items():
            if df_col_str == excel_col_str:
                mapping[df_col] = excel_col_num
                break

    return mapping


def extract_section_from_course(course_name):
    if not course_name or course_name in ["Free Slot", "Free Slot (Lab)", "Free Slot (Class)"]:
        return None, course_name
    section_patterns = [
        r'\(([A-Z]{2,3}-[A-Z])\)',
        r'\(([A-Z]{2,3},\s*\d{2,4})\)',
        r'\(([A-Z]{2,3})\)',
    ]
    for pattern in section_patterns:
        match = re.search(pattern, course_name)
        if match:
            section = match.group(1)
            clean_course_name = re.sub(pattern, '', course_name).strip()
            return section, clean_course_name
    return None, course_name

def parse_course_and_time(raw_course, column_time):
    column_time = normalize_time_str(column_time)
    if pd.isna(raw_course) or str(raw_course).strip() == "":
        return column_time, "Free Slot", None, None
    course_text = str(raw_course).strip()
    if course_text.upper() == "FSM":
        return column_time, "FSM", None, None
    time_pattern = re.compile(r'(\d{1,2}:\d{2}\s*[-–]\s*\d{1,2}:\d{2})')
    time_matches = time_pattern.findall(course_text)
    if time_matches:
        actual_time = normalize_time_str(time_matches[0])
        clean_course = time_pattern.sub('', course_text).strip()
    else:
        actual_time = column_time
        clean_course = course_text
    section, final_course_name = extract_section_from_course(clean_course)
    final_course_name = re.sub(r'\s+', ' ', final_course_name).strip()
    if not final_course_name:
        return actual_time, "Free Slot", None, None
    return actual_time, final_course_name, section, clean_course

def process_lab_section(df, day_name, color_batch_map, ws, header_row_excel):
    lab_start_idx = df[df.iloc[:, 0].astype(str).str.contains("Lab", case=False, na=False)].index
    if len(lab_start_idx) == 0:
        return pd.DataFrame()
    lab_start_idx = lab_start_idx[0]
    lab_df = df.iloc[lab_start_idx:].copy()
    time_slots_row = lab_df.iloc[0]
    time_slot_columns = {}
    for col_idx in range(len(time_slots_row)):
        slot = time_slots_row.iloc[col_idx]
        if pd.notna(slot) and str(slot).strip():
            slot_str = normalize_time_str(str(slot).strip())
            if ":" in slot_str or "break" in slot_str.lower():
                time_slot_columns[slot_str] = col_idx
    if not time_slot_columns:
        return pd.DataFrame()
    excel_lab_start_row = header_row_excel + lab_start_idx - 2
    results = []
    for pandas_idx, (r_idx, row) in enumerate(lab_df.iloc[1:].iterrows()):
        room = row.iloc[0]
        if pd.isna(room) or str(room).strip() == "" or str(room).strip() == "Lab":
            continue
        excel_row = excel_lab_start_row + pandas_idx
        for time_slot, col_idx in time_slot_columns.items():
            course = row.iloc[col_idx] if col_idx < len(row) else None
            actual_time, course_name, section, original_course = parse_course_and_time(course, time_slot)
            if course_name == "Free Slot":
                final_course_name = "Free Slot (Lab)"
                section = None
                batch = None
            elif course_name.upper() == "FSM":
                final_course_name = "FSM"
                section = None
                excel_col = col_idx + 1
                try:
                    cell_color = normalize_color(ws.cell(row=excel_row, column=excel_col).fill.fgColor)
                    batch = color_batch_map.get(cell_color, None)
                except:
                    batch = None
            else:
                final_course_name = course_name
                excel_col = col_idx + 1
                try:
                    cell_color = normalize_color(ws.cell(row=excel_row, column=excel_col).fill.fgColor)
                    batch = color_batch_map.get(cell_color, None)
                except:
                    batch = None
            if (final_course_name not in ["Free Slot (Lab)", "FSM"] and 
                "lab" not in final_course_name.lower()):
                final_course_name += " Lab"
            results.append({
                "Day": day_name,
                "Course Name": final_course_name,
                "Class Time": actual_time,
                "Room No": room,
                "Section": section,
                "Batch": batch,
                "Type": "Lab",
            })
    return pd.DataFrame(results)

def reshape_timetable(df, day_name):
    original_df = df.copy()
    df = df.dropna(how="all").dropna(axis=1, how="all").reset_index(drop=True)
    header_row_index = df.index[df.iloc[:, 0].astype(str).str.contains("Room", case=False, na=False)]
    if len(header_row_index) == 0:
        return pd.DataFrame(columns=["Day", "Course Name", "Class Time", "Room No", "Section", "Batch", "Type"])
    header_row_pandas = header_row_index[0]
    wb = load_workbook(file, data_only=False)
    ws = wb[day_name]
    header_row_excel = None
    for row_num in range(1, 20):
        cell_value = ws.cell(row=row_num, column=1).value
        if cell_value and "room" in str(cell_value).lower():
            header_row_excel = row_num
            break
    if header_row_excel is None:
        return pd.DataFrame(columns=["Day", "Course Name", "Class Time", "Room No", "Section", "Batch", "Type"])
    color_batch_map = extract_color_batch_map(file, day_name)
    df.columns = df.iloc[header_row_pandas]
    # Fill forward merged/blank headers so times like 9:00-10:45 aren't lost
    df.columns = df.columns.fillna(method='ffill')
    df.columns = [normalize_time_str(str(c)) if pd.notna(c) else c for c in df.columns]
    df = df.iloc[header_row_pandas + 1:].reset_index(drop=True)
    excel_df_mapping = create_excel_to_dataframe_mapping(ws, df.columns, header_row_excel)
    lab_start_idx = original_df[original_df.iloc[:, 0].astype(str).str.contains("Lab", case=False, na=False)].index
    if len(lab_start_idx) > 0:
        lab_start_idx = lab_start_idx[0]
        class_df = df.iloc[:lab_start_idx - header_row_pandas - 1]
    else:
        class_df = df
    time_pattern = re.compile(r"\d{1,2}:\d{2}\s*[-–]\s*\d{1,2}:\d{2}")
    time_columns = [col for col in class_df.columns if pd.notna(col) and time_pattern.search(str(col))]
    time_columns = [normalize_time_str(col) for col in time_columns]
    results = []
    for r_idx, row in class_df.iterrows():
        room = row.get("Room", "Unknown")
        if pd.isna(room) or "Lab" in str(room):
            continue
        for time_col in time_columns:
            raw_course = row.get(time_col, "") if time_col in row else ""
            # detect times inside course cells even if header doesn't match
            match_inside = re.search(r"\d{1,2}:\d{2}\s*[-–]\s*\d{1,2}:\d{2}", str(raw_course))
            if match_inside:
                time_val = normalize_time_str(match_inside.group(0))
            else:
                time_val = time_col
            actual_time, course_name, section, original_course = parse_course_and_time(raw_course, time_val)

            if course_name == "Free Slot" and str(raw_course).strip() != "":
                pass
            elif course_name == "Free Slot":
                course_name = "Free Slot (Class)"
            batch = None
            excel_col = excel_df_mapping.get(time_col)
            if excel_col is not None:
                excel_row = header_row_excel + 1 + r_idx
                try:
                    cell_color = normalize_color(ws.cell(row=excel_row, column=excel_col).fill.fgColor)
                    batch = color_batch_map.get(cell_color, None)
                except:
                    pass
            entry = {
                "Day": day_name,
                "Course Name": course_name,
                "Class Time": actual_time,
                "Room No": room,
                "Section": section,
                "Batch": batch,
                "Type": "Class",
            }
            results.append(entry)
    lab_results = process_lab_section(original_df, day_name, color_batch_map, ws, header_row_excel)
    final_df = pd.DataFrame(results)
    if not lab_results.empty:
        final_df = pd.concat([final_df, lab_results], ignore_index=True)
    return final_df

def get_time_table():
    sheet_names = pd.ExcelFile(file).sheet_names
    time_table_data = {}
    for names in sheet_names:
        if names != "Welcome":
            data = pd.read_excel(file, sheet_name=names)
            time_table_data[names] = data
    event_tables = {day: reshape_timetable(time_table_data[day], day) for day in time_table_data}
    all_days_df = pd.concat(event_tables.values())
    unwanted_slots = ["05:20-06:40", "06:45-08:05", "05:20-08:05"]
    all_days_df = all_days_df[~all_days_df["Class Time"].isin(unwanted_slots)]
    return all_days_df
